from flask import render_template, request, redirect, url_for, flash, session, abort, make_response, send_file, send_from_directory
from werkzeug.security import generate_password_hash
from flask_login import current_user
from werkzeug.utils import secure_filename
from sqlalchemy import extract, and_
from app import app, db
from models import User, Ticket, TicketComment, Attachment
from forms import LoginForm, TicketForm, UpdateTicketForm, CommentForm, UserRegistrationForm, AssignTicketForm, UserProfileForm
from datetime import datetime
from utils.email import send_assignment_email  # Add this import
from utils.timezone import utc_to_ist
import logging
import os
import socket
import platform
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'bmp', 'pdf', 'doc', 'docx', 'xls', 'xlsx', 'csv', 'ppt', 'pptx'}
UPLOAD_FOLDER = 'uploads/'  # Set a secure uploads folder

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# Helper function to check if user is logged in
def is_logged_in():
    return 'user_id' in session

# Helper function to get current user
def get_current_user():
    if is_logged_in():
        return User.query.get(session['user_id'])
    return None

# Helper function to require login
def login_required(f):
    def decorated_function(*args, **kwargs):
        if not is_logged_in():
            flash('Please log in to access this page.', 'warning')
            return redirect(url_for('common_login'))
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

# Helper function to require super admin
def super_admin_required(f):
    def decorated_function(*args, **kwargs):
        if not is_logged_in():
            flash('Please log in to access this page.', 'warning')
            return redirect(url_for('common_login'))
        user = get_current_user()
        if not user or not user.is_super_admin:
            flash('Super Admin access required.', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

@app.route('/')
def index():
    """Home page"""
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def common_login():
    """Common login page for all user types"""
    if is_logged_in():
        user = get_current_user()
        # Redirect to appropriate dashboard based on role
        if user.is_super_admin:
            return redirect(url_for('super_admin_dashboard'))
        else:
            return redirect(url_for('user_dashboard'))
    
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user and user.check_password(form.password.data):
            # Set session variables
            session['user_id'] = user.id
            session['role'] = user.role
            
            # Update IP address and system info
            user.ip_address = request.environ.get('HTTP_X_FORWARDED_FOR', request.environ.get('REMOTE_ADDR'))
            try:
                user.system_name = socket.gethostname()
            except:
                user.system_name = 'Unknown'
            db.session.commit()
            
            flash(f'Welcome back, {user.first_name}!', 'success')
            
            # Route to appropriate dashboard based on role
            if user.is_super_admin:
                return redirect(url_for('super_admin_dashboard'))
            else:
                return redirect(url_for('user_dashboard'))
        else:
            flash('Invalid username or password.', 'error')
    
    return render_template('common_login.html', form=form)

@app.route('/user-login', methods=['GET', 'POST'])
def user_login():
    """User login page - redirects to common login"""
    return redirect(url_for('common_login'))

@app.route('/admin-login', methods=['GET', 'POST'])
def admin_login():
    """Admin login page - redirects to common login"""
    return redirect(url_for('common_login'))

@app.route('/logout')
def logout():
    """Logout user"""
    session.clear()
    flash('You have been logged out.', 'info')
    return redirect(url_for('index'))

@app.route('/user-dashboard')
@login_required
def user_dashboard():
    """User dashboard showing their tickets"""
    user = get_current_user()
    
    # Get filter parameters
    status_filter = request.args.get('status', 'all')
    search_query = request.args.get('search', '')
    
    # Build query
    query = Ticket.query.filter_by(user_id=user.id)
    
    if status_filter != 'all':
        query = query.filter_by(status=status_filter)
    
    if search_query:
        query = query.filter(Ticket.title.contains(search_query))
    
    tickets = query.order_by(Ticket.created_at.desc()).all()
    
    return render_template('user_dashboard.html', user=user, tickets=tickets, 
                         status_filter=status_filter, search_query=search_query)

@app.route('/user-profile', methods=['GET', 'POST'])
@login_required
def user_profile():
    """User profile management"""
    user = get_current_user()
    form = UserProfileForm(obj=user)
    
    if form.validate_on_submit():
        user.first_name = form.first_name.data
        user.last_name = form.last_name.data
        user.email = form.email.data
        user.department = form.department.data
        user.system_name = form.system_name.data
        
        # Handle profile image upload
        if 'profile_image' in request.files:
            file = request.files['profile_image']
            if file.filename != '':
                filename = secure_filename(file.filename)
                # Save file logic would go here
                user.profile_image = filename
        
        db.session.commit()
        flash('Profile updated successfully!', 'success')
        return redirect(url_for('user_profile'))
    
    return render_template('user_profile.html', form=form, user=user)

@app.route('/super-admin-dashboard')
@super_admin_required
def super_admin_dashboard():
    """Super Admin dashboard with full system overview and filters"""
    user = get_current_user()
    if not user.is_super_admin:
        flash('Super Admin access required.', 'error')
        return redirect(url_for('index'))

    # Comprehensive statistics (unfiltered)
    total_tickets = Ticket.query.count()
    open_tickets = Ticket.query.filter_by(status='Open').count()
    in_progress_tickets = Ticket.query.filter_by(status='In Progress').count()
    resolved_tickets = Ticket.query.filter_by(status='Resolved').count()
    total_users = User.query.filter_by(role='user').count()
    total_admins = User.query.filter_by(role='admin').count()
    hardware_tickets = Ticket.query.filter_by(category='Hardware').count()
    software_tickets = Ticket.query.filter_by(category='Software').count()

    stats = {
        'total_tickets': total_tickets,
        'open_tickets': open_tickets,
        'in_progress_tickets': in_progress_tickets,
        'resolved_tickets': resolved_tickets,
        'total_users': total_users,
        'total_admins': total_admins,
        'hardware_tickets': hardware_tickets,
        'software_tickets': software_tickets
    }

    # Filter parameters for recent tickets
    status_filter = request.args.get('status', 'all')
    priority_filter = request.args.get('priority', 'all')
    category_filter = request.args.get('category', 'all')
    search_query = request.args.get('search', '')
    day_filter = request.args.get('day', '')
    month_filter = request.args.get('month', '')
    year_filter = request.args.get('year', '')

    # Build filtered query for recent tickets
    query = Ticket.query
    if status_filter != 'all':
        query = query.filter_by(status=status_filter)
    if priority_filter != 'all':
        query = query.filter_by(priority=priority_filter)
    if category_filter != 'all':
        query = query.filter_by(category=category_filter)
    if search_query:
        query = query.filter(Ticket.title.contains(search_query))
    if year_filter:
        query = query.filter(extract('year', Ticket.created_at) == int(year_filter))
    if month_filter:
        query = query.filter(extract('month', Ticket.created_at) == int(month_filter))
    if day_filter:
        query = query.filter(extract('day', Ticket.created_at) == int(day_filter))

    recent_tickets = query.order_by(Ticket.created_at.desc()).limit(10).all()

    return render_template(
        'super_admin_dashboard.html',
        stats=stats,
        recent_tickets=recent_tickets,
        status_filter=status_filter,
        priority_filter=priority_filter,
        category_filter=category_filter,
        search_query=search_query,
        day_filter=day_filter,
        month_filter=month_filter,
        year_filter=year_filter
    )




@app.route('/create-ticket', methods=['GET', 'POST'])
@login_required
def create_ticket():
    """Create a new ticket"""
    form = TicketForm()
    user = get_current_user()

    if form.validate_on_submit():
        # Capture current session IP and system info for this specific ticket
        current_ip = request.environ.get('HTTP_X_FORWARDED_FOR',
                                         request.environ.get('HTTP_X_REAL_IP',
                                                             request.environ.get('REMOTE_ADDR')))

        # Get system name
        current_system_name = form.system_name.data.strip() if form.system_name.data and form.system_name.data.strip() else None
        if not current_system_name:
            if user.system_name and user.system_name.strip():
                current_system_name = user.system_name.strip()
            else:
                user_agent = request.headers.get('User-Agent', '').lower()
                if 'windows' in user_agent:
                    current_system_name = 'Windows System'
                elif 'mac os x' in user_agent or 'macos' in user_agent:
                    current_system_name = 'macOS System'
                elif 'linux' in user_agent:
                    current_system_name = 'Linux System'
                elif 'android' in user_agent:
                    current_system_name = 'Android Device'
                elif 'iphone' in user_agent or 'ipad' in user_agent:
                    current_system_name = 'iOS Device'
                else:
                    current_system_name = f'Unknown System ({request.remote_addr})'

        user.ip_address = current_ip
        user.system_name = current_system_name

        # Handle file uploads (supporting multiple attachments)
        attachment_filenames = []
        files = request.files.getlist('attachments')
        for file in files:
            if file and file.filename and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_')
                unique_filename = timestamp + filename
                os.makedirs(UPLOAD_FOLDER, exist_ok=True)
                filepath = os.path.join(UPLOAD_FOLDER, unique_filename)
                try:
                    file.save(filepath)
                    attachment_filenames.append(unique_filename)
                except Exception as e:
                    flash(f'Error uploading file {filename}.', 'warning')

        # For backward compatibility, store the first image filename in image_filename, others in attachments
        image_filename = None
        other_attachments = []
        for fname in attachment_filenames:
            if any(fname.lower().endswith(ext) for ext in
                   ['.png', '.jpg', '.jpeg', '.gif', '.bmp']) and not image_filename:
                image_filename = fname
            else:
                other_attachments.append(fname)

        # Create the ticket first
        ticket = Ticket(
            title=form.title.data,
            description=form.description.data,
            category=form.category.data,
            priority=form.priority.data,
            user_id=user.id,
            user_name=user.full_name,
            user_ip_address=current_ip,
            user_system_name=current_system_name,
            image_filename=image_filename
        )
        db.session.add(ticket)
        db.session.commit()

        # Create attachment records for non-image files
        for attachment_filename in other_attachments:
            attachment = Attachment(
                ticket_id=ticket.id,
                filename=attachment_filename
            )
            db.session.add(attachment)
        
        if other_attachments:
            db.session.commit()

        flash(f'Ticket {ticket.ticket_number} created successfully!', 'success')
        return redirect(url_for('user_dashboard'))

    return render_template('create_ticket.html', form=form)

@app.route('/ticket/<int:ticket_id>')
@login_required
def view_ticket(ticket_id):
    """View ticket details"""
    ticket = Ticket.query.get_or_404(ticket_id)
    user = get_current_user()
    
    # Check if user can view this ticket
    if not user.is_super_admin and ticket.user_id != user.id:
        abort(403)
    
    form = CommentForm()
    assign_form = AssignTicketForm() if user.is_super_admin else None
    
    return render_template('view_ticket.html', ticket=ticket, form=form, 
                         assign_form=assign_form, user=user)

@app.route('/ticket/<int:ticket_id>/comment', methods=['POST'])
@login_required
def add_comment(ticket_id):
    """Add comment to ticket"""
    ticket = Ticket.query.get_or_404(ticket_id)
    user = get_current_user()
    
    # Check if user can comment on this ticket
    if not user.is_super_admin and ticket.user_id != user.id:
        abort(403)
    
    form = CommentForm()
    if form.validate_on_submit():
        comment = TicketComment(
            ticket_id=ticket_id,
            user_id=user.id,
            comment=form.comment.data
        )
        db.session.add(comment)
        ticket.updated_at = datetime.utcnow()
        db.session.commit()
        
        flash('Comment added successfully!', 'success')
    
    return redirect(url_for('view_ticket', ticket_id=ticket_id))

@app.route('/ticket/<int:ticket_id>/edit', methods=['GET', 'POST'])
@super_admin_required
def edit_ticket(ticket_id):
    """Edit ticket (admin only)"""
    ticket = Ticket.query.get_or_404(ticket_id)
    form = UpdateTicketForm(obj=ticket)
    current_user = get_current_user()
    
    if form.validate_on_submit():
        # Only status can be updated - no one can edit title, description, category, or priority
        old_status = ticket.status
        ticket.status = form.status.data
        
        # Set resolved_at if status changed to Resolved
        if old_status != 'Resolved' and ticket.status == 'Resolved':
            ticket.resolved_at = datetime.utcnow()
        elif ticket.status != 'Resolved':
            ticket.resolved_at = None
        
        ticket.updated_at = datetime.utcnow()
        
        # Add comment if super admin provided one
        admin_comment = request.form.get('admin_comment', '').strip()
        if current_user and current_user.is_super_admin and admin_comment:
            comment = TicketComment(
                ticket_id=ticket.id,
                user_id=current_user.id,
                comment=f"Status updated to '{ticket.status}'. {admin_comment}"
            )
            db.session.add(comment)
        elif old_status != ticket.status:
            # Add automatic status change comment
            comment = TicketComment(
                ticket_id=ticket.id,
                user_id=current_user.id,
                comment=f"Status updated from '{old_status}' to '{ticket.status}'"
            )
            db.session.add(comment)
        
        db.session.commit()
        
        flash('Ticket status updated successfully!', 'success')
        return redirect(url_for('view_ticket', ticket_id=ticket_id))
    
    return render_template('edit_ticket.html', form=form, ticket=ticket, user=current_user)


@app.route('/ticket/<int:ticket_id>/assign', methods=['POST'])
@super_admin_required
def assign_ticket(ticket_id):
    """Assign ticket to admin"""
    ticket = Ticket.query.get_or_404(ticket_id)
    form = AssignTicketForm()

    if form.validate_on_submit():
        current_user = get_current_user()
        ticket.assigned_to = form.assigned_to.data
        ticket.assigned_by = current_user.id if current_user else None
        if ticket.status == 'Open':
            ticket.status = 'In Progress'
        ticket.updated_at = datetime.utcnow()
        ticket.assigned_at = datetime.utcnow()
        db.session.commit()

        assignee = User.query.get(form.assigned_to.data)

        # Send email notification
        if assignee and assignee.email:
            send_assignment_email(assignee.email, ticket.id, assignee.full_name)

        flash(f'Ticket assigned to {assignee.full_name}!', 'success')

    return redirect(url_for('view_ticket', ticket_id=ticket_id))


@app.route('/edit-user/<int:user_id>', methods=['GET', 'POST'])
@super_admin_required
def edit_user(user_id):
    """Edit user (Super Admin only)"""
    current_user = get_current_user()
    if not current_user.is_super_admin:
        flash('Super Admin access required.', 'error')
        return redirect(url_for('index'))

    user = User.query.get_or_404(user_id)
    form = UserProfileForm()

    if form.validate_on_submit():
        user.username = form.username.data
        user.role = form.role.data
        user.first_name = form.first_name.data
        user.last_name = form.last_name.data
        user.email = form.email.data
        user.department = form.department.data
        user.system_name = form.system_name.data
        # Only update password if a new value is provided
        if form.password.data:
            user.password = generate_password_hash(form.password.data)
        db.session.commit()
        flash(f'User {user.username} updated successfully!', 'success')
        return redirect(url_for('view_user', user_id=user_id))

    # Pre-populate the form
    form.username.data = user.username
    form.role.data = user.role
    form.first_name.data = user.first_name
    form.last_name.data = user.last_name
    form.email.data = user.email
    form.department.data = user.department
    form.system_name.data = user.system_name
    # Do not pre-fill password for security reasons

    return render_template('edit_user.html', form=form, user=user)

@app.route('/manage-users')
@super_admin_required
def manage_users():
    """Super Admin user management"""
    user = get_current_user()
    if not user.is_super_admin:
        flash('Super Admin access required.', 'error')
        return redirect(url_for('index'))
    
    users = User.query.all()
    return render_template('manage_users.html', users=users)

@app.route('/create-user', methods=['GET', 'POST'])
@super_admin_required
def create_user():
    """Create new user (Super Admin only)"""
    user = get_current_user()
    if not user.is_super_admin:
        flash('Super Admin access required.', 'error')
        return redirect(url_for('index'))
    
    form = UserRegistrationForm()
    if form.validate_on_submit():
        new_user = User(
            username=form.username.data,
            email=form.email.data,
            first_name=form.first_name.data,
            last_name=form.last_name.data,
            department=form.department.data,
            role=form.role.data
        )
        new_user.set_password(form.password.data)
        db.session.add(new_user)
        db.session.commit()
        
        flash(f'User {new_user.username} created successfully!', 'success')
        return redirect(url_for('manage_users'))
    
    return render_template('create_user.html', form=form)

@app.route('/view-user/<int:user_id>')
@super_admin_required
def view_user(user_id):
    """View user details (Super Admin only)"""
    current_user = get_current_user()
    if not current_user.is_super_admin:
        flash('Super Admin access required.', 'error')
        return redirect(url_for('index'))
    
    user = User.query.get_or_404(user_id)
    # Get user's tickets
    user_tickets = Ticket.query.filter_by(user_id=user_id).order_by(Ticket.created_at.desc()).all()
    assigned_tickets = Ticket.query.filter_by(assigned_to=user_id).order_by(Ticket.created_at.desc()).all()
    
    return render_template('view_user.html', user=user, user_tickets=user_tickets, assigned_tickets=assigned_tickets)



@app.route('/delete-user/<int:user_id>', methods=['POST'])
@super_admin_required
def delete_user(user_id):
    """Delete user (Super Admin only)"""
    current_user = get_current_user()
    if not current_user.is_super_admin:
        flash('Super Admin access required.', 'error')
        return redirect(url_for('index'))
    
    user_to_delete = User.query.get_or_404(user_id)
    
    # Prevent deletion of super admin users
    if user_to_delete.role == 'super_admin':
        flash('Cannot delete Super Admin users for security reasons.', 'error')
        return redirect(url_for('manage_users'))
    
    # Prevent self-deletion
    if user_to_delete.id == current_user.id:
        flash('You cannot delete your own account.', 'error')
        return redirect(url_for('manage_users'))
    
    try:
        # Get user's tickets and reassign or handle them
        user_tickets = Ticket.query.filter_by(user_id=user_id).all()
        assigned_tickets = Ticket.query.filter_by(assigned_to=user_id).all()
        
        # Update tickets created by this user to preserve data integrity
        for ticket in user_tickets:
            ticket.user_name = f"{user_to_delete.full_name} (Deleted User)"
            ticket.user_id = None  # Set to None to indicate deleted user
        
        # Reassign tickets that were assigned to this user
        for ticket in assigned_tickets:
            ticket.assigned_to = None
            ticket.status = 'Open'  # Reset status to Open for reassignment
        
        # Delete user's comments (cascade should handle this, but being explicit)
        comments = TicketComment.query.filter_by(user_id=user_id).all()
        for comment in comments:
            db.session.delete(comment)
        
        # Delete the user
        username = user_to_delete.username
        db.session.delete(user_to_delete)
        db.session.commit()
        
        flash(f'User "{username}" has been successfully deleted. Their tickets have been preserved and reassigned tickets are now available for assignment.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting user: {str(e)}', 'error')
    
    return redirect(url_for('manage_users'))

@app.route('/assign-work/<int:ticket_id>', methods=['GET', 'POST'])
@super_admin_required
def assign_work(ticket_id):
    """Super Admin assigns work to specific admins based on category"""
    user = get_current_user()
    if not user.is_super_admin:
        flash('Super Admin access required.', 'error')
        return redirect(url_for('index'))
    
    ticket = Ticket.query.get_or_404(ticket_id)
    
    # Get all super admins for assignment (simplified role structure)
    admins = User.query.filter_by(role='super_admin').all()
    
    form = AssignTicketForm()
    form.assigned_to.choices = [(admin.id, f"{admin.full_name} ({admin.department})") for admin in admins]
    
    if form.validate_on_submit():
        ticket.assigned_to = form.assigned_to.data
        ticket.assigned_by = user.id
        ticket.status = 'In Progress'
        ticket.updated_at = datetime.utcnow()
        db.session.commit()
        
        assignee = User.query.get(form.assigned_to.data)
        flash(f'Work assigned to {assignee.full_name}!', 'success')
        return redirect(url_for('super_admin_dashboard'))
    
    return render_template('assign_work.html', form=form, ticket=ticket, admins=admins)

def create_default_admin():
    """Create default super admin and test user if none exists"""
    try:
        super_admin = User.query.filter_by(role='super_admin').first()
        if not super_admin:
            # Create Super Admin
            super_admin_user = User(
                username='superadmin',
                email='superadmin@gtnengineering.com',
                first_name='Super',
                last_name='Administrator',
                department='IT',
                role='super_admin'
            )
            super_admin_user.set_password('super123')
            db.session.add(super_admin_user)
            db.session.commit()
            
            # Create a test user
            test_user = User(
                username='testuser',
                email='user@gtnengineering.com',
                first_name='Test',
                last_name='User',
                department='Engineering',
                role='user'
            )
            test_user.set_password('test123')
            db.session.add(test_user)
            db.session.commit()
            
            logging.info("Default super admin and test user created")
    except Exception as e:
        logging.error(f"Error creating default users: {e}")
        db.session.rollback()

@app.route('/reports-dashboard')
@super_admin_required
def reports_dashboard():
    """Reports Dashboard with visual analytics (Super Admin only)"""
    current_user = get_current_user()
    if not current_user.is_super_admin:
        flash('Super Admin access required.', 'error')
        return redirect(url_for('index'))
    
    # Get comprehensive statistics
    total_tickets = Ticket.query.count()
    open_tickets = Ticket.query.filter_by(status='Open').count()
    in_progress_tickets = Ticket.query.filter_by(status='In Progress').count()
    resolved_tickets = Ticket.query.filter_by(status='Resolved').count()
    closed_tickets = Ticket.query.filter_by(status='Closed').count()
    
    # Category breakdown
    hardware_tickets = Ticket.query.filter_by(category='Hardware').count()
    software_tickets = Ticket.query.filter_by(category='Software').count()
    network_tickets = Ticket.query.filter_by(category='Network').count()
    other_tickets = Ticket.query.filter_by(category='Other').count()
    
    # Priority breakdown  
    critical_tickets = Ticket.query.filter_by(priority='Critical').count()
    high_tickets = Ticket.query.filter_by(priority='High').count()
    medium_tickets = Ticket.query.filter_by(priority='Medium').count()
    low_tickets = Ticket.query.filter_by(priority='Low').count()
    
    # Get all tickets for detailed table
    all_tickets = Ticket.query.order_by(Ticket.created_at.desc()).all()
    
    stats = {
        'total_tickets': total_tickets,
        'open_tickets': open_tickets,
        'in_progress_tickets': in_progress_tickets,
        'resolved_tickets': resolved_tickets,
        'closed_tickets': closed_tickets,
        'hardware_tickets': hardware_tickets,
        'software_tickets': software_tickets,
        'network_tickets': network_tickets,
        'other_tickets': other_tickets,
        'critical_tickets': critical_tickets,
        'high_tickets': high_tickets,
        'medium_tickets': medium_tickets,
        'low_tickets': low_tickets
    }
    
    # Prepare chart data for JavaScript
    chart_data = {
        'category': [hardware_tickets, software_tickets, network_tickets, other_tickets],
        'priority': [critical_tickets, high_tickets, medium_tickets, low_tickets],
        'status': [open_tickets, in_progress_tickets, resolved_tickets, closed_tickets]
    }
    
    return render_template('reports_dashboard.html', stats=stats, tickets=all_tickets, chart_data=chart_data)

@app.route('/edit-assignment/<int:ticket_id>', methods=['GET', 'POST'])
@super_admin_required
def edit_assignment(ticket_id):
    """Edit ticket assignment (Super Admin only)"""
    current_user = get_current_user()
    if not current_user.is_super_admin:
        flash('Super Admin access required.', 'error')
        return redirect(url_for('index'))
    
    ticket = Ticket.query.get_or_404(ticket_id)
    
    if request.method == 'POST':
        assigned_to = request.form.get('assigned_to')
        if assigned_to:
            assigned_to = int(assigned_to) if assigned_to != '0' else None
        else:
            assigned_to = None
            
        ticket.assigned_to = assigned_to
        ticket.updated_at = datetime.utcnow()
        
        try:
            db.session.commit()
            assignee_name = User.query.get(assigned_to).full_name if assigned_to else 'Unassigned'
            flash(f'Ticket {ticket.ticket_number} has been assigned to {assignee_name}.', 'success')
            return redirect(url_for('super_admin_dashboard'))
        except Exception as e:
            db.session.rollback()
            flash('Error updating assignment. Please try again.', 'error')
            logging.error(f"Error updating ticket assignment: {e}")
    
    # Get all admin users for assignment dropdown
    admin_users = User.query.filter_by(role="super_admin").all()
    
    return render_template('edit_assignment.html', ticket=ticket, admin_users=admin_users)

@app.route('/view-image/<filename>')
@login_required
def view_image(filename):
    """View uploaded ticket image - admins can view any, users can view their own"""
    current_user = get_current_user()
    
    # Find ticket with this image
    ticket = Ticket.query.filter_by(image_filename=filename).first()
    if not ticket:
        abort(404)
    
    # Check permissions - admins can view any, users only their own tickets
    if not current_user.is_super_admin and ticket.user_id != current_user.id:
        abort(403)
    
    try:
        return send_from_directory('uploads', filename)
    except FileNotFoundError:
        abort(404)

@app.route('/download-attachment/<filename>')
@login_required
def download_attachment(filename):
    """Download file attachment - admins can download any, users can download their own"""
    current_user = get_current_user()
    
    # Find the attachment record
    attachment = Attachment.query.filter_by(filename=filename).first()
    if not attachment:
        abort(404)
    
    # Check permissions - admins can download any, users only their own tickets
    if not current_user.is_super_admin:
        ticket = Ticket.query.get(attachment.ticket_id)
        if not ticket or ticket.user_id != current_user.id:
            abort(403)
    
    try:
        return send_from_directory('uploads', filename, as_attachment=True)
    except FileNotFoundError:
        abort(404)

@app.route('/download-excel-report')
@super_admin_required
def download_excel_report():
    """Download Excel report of all tickets (Super Admin only) with filtering options"""
    current_user = get_current_user()
    if not current_user.is_super_admin:
        flash('Super Admin access required.', 'error')
        return redirect(url_for('index'))

    try:
        # --- FILTER PARAMS ---
        filter_mode = request.args.get('filter_mode', 'range')
        from_date = request.args.get('from_date')
        to_date = request.args.get('to_date')
        month = request.args.get('month')
        year = request.args.get('year')

        # --- BUILD QUERY BASED ON FILTER ---
        query = Ticket.query.join(User, Ticket.user_id == User.id)

        if filter_mode == 'range' and from_date and to_date:
            from_dt = datetime.strptime(from_date, '%Y-%m-%d')
            to_dt = datetime.strptime(to_date, '%Y-%m-%d')
            query = query.filter(Ticket.created_at >= from_dt, Ticket.created_at <= to_dt)
        elif filter_mode == 'month' and month:
            y, m = map(int, month.split('-'))
            query = query.filter(
                extract('year', Ticket.created_at) == y,
                extract('month', Ticket.created_at) == m
            )
        elif filter_mode == 'year' and year:
            query = query.filter(extract('year', Ticket.created_at) == int(year))

        tickets = query.all()

        # --- EXCEL GENERATION ---
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tickets Report"
        headers = [
            'Ticket ID', 'Title', 'Description', 'Category', 'Priority', 'Status',
            'Created By', 'User Email', 'User Department', 'System Name', 'IP Address',
            'Assigned To', 'Assigned By', 'Created Date', 'Updated Date', 'Resolved Date'
        ]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Add headers to worksheet
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Add ticket data
        for row, ticket in enumerate(tickets, 2):
            assignee_name = ticket.assignee.full_name if hasattr(ticket, 'assignee') and ticket.assignee else 'Unassigned'
            assigner_name = ticket.assigner.full_name if hasattr(ticket, 'assigner') and ticket.assigner else 'N/A'
            data = [
                ticket.ticket_number,
                ticket.title,
                ticket.description,
                ticket.category,
                ticket.priority,
                ticket.status,
                getattr(ticket, 'user_name', ticket.user.full_name if hasattr(ticket.user, 'full_name') else 'N/A'),
                ticket.user.email,
                getattr(ticket.user, 'department', 'N/A') or 'N/A',
                getattr(ticket, 'user_system_name', 'N/A') or 'N/A',
                getattr(ticket, 'user_ip_address', 'N/A') or 'N/A',
                assignee_name,
                assigner_name,
                utc_to_ist(ticket.created_at).strftime('%Y-%m-%d %H:%M:%S') if ticket.created_at else 'N/A',
                utc_to_ist(ticket.updated_at).strftime('%Y-%m-%d %H:%M:%S') if ticket.updated_at else 'N/A',
                utc_to_ist(ticket.resolved_at).strftime('%Y-%m-%d %H:%M:%S') if ticket.resolved_at else 'N/A'
            ]
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'GTN_Helpdesk_Report_{timestamp}.xlsx'

        # Create response
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'

        return response

    except Exception as e:
        logging.error(f"Error generating Excel report: {e}")
        flash('Error generating report. Please try again.', 'error')
        return redirect(url_for('reports_dashboard'))

# Initialize default admin on first import
with app.app_context():
    create_default_admin()

# Error handlers
@app.errorhandler(404)
def not_found_error(error):
    return render_template('404.html'), 404

@app.errorhandler(403)
def forbidden_error(error):
    return render_template('403.html'), 403

@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    return render_template('500.html'), 500
